import os
import io
import pandas as pd
from openpyxl.styles import Border, PatternFill, Font

DRIVE_ROOT = "/content/drive/Shareddrives/SMART TECH TANK - DATA DEPARTMENT/2026"

MAIN_RACES = ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]
DM_EXTRA_RACES = ["MELAYU", "CINA", "INDIA"]
PARTY_COLS = ["PAS", "PKR", "PPBM", "UMNO"]
AGE_GROUPS = ["18-24", "25-30", "31-40", "41-50", "51-60", "61+"]


def list_dirs(path):
    try:
        return sorted([
            f for f in os.listdir(path)
            if os.path.isdir(os.path.join(path, f))
        ])
    except Exception:
        return []


def list_excel_files(path):
    try:
        return sorted([
            f for f in os.listdir(path)
            if f.lower().endswith((".xlsx", ".xls"))
            and not f.startswith("~$")
        ])
    except Exception:
        return []


def normalise_race(race):
    r = str(race).strip().upper()

    if r in {"MELAYU", "CINA", "INDIA"}:
        return r

    return "LAIN-LAIN"


def get_age_group(age):
    try:
        a = int(float(age))

        if 18 <= a <= 24:
            return "18-24"
        elif 25 <= a <= 30:
            return "25-30"
        elif 31 <= a <= 40:
            return "31-40"
        elif 41 <= a <= 50:
            return "41-50"
        elif 51 <= a <= 60:
            return "51-60"
        elif a >= 61:
            return "61+"

    except Exception:
        pass

    return None


def get_age_num(age):
    try:
        return int(float(age))
    except Exception:
        return None


def pct(part, total):
    return round(part / total * 100, 2) if total else 0


def build_row(base, grp):
    total = grp["nokp"].notna().sum()

    race_vc = grp["race_norm"].value_counts()
    age_vc = grp["age_group"].value_counts()
    sex_vc = grp["jantina"].astype(str).str.strip().str.upper().value_counts()
    party_vc = grp["party"].astype(str).str.strip().str.upper().value_counts()

    row = {**base, "JUMLAH": total}

    for s, label in [("L", "LELAKI"), ("P", "PEREMPUAN")]:
        c = sex_vc.get(s, 0)
        row[label] = c
        row[f"{label} (%)"] = pct(c, total)

    for r in MAIN_RACES:
        c = race_vc.get(r, 0)
        row[r] = c
        row[f"{r} (%)"] = pct(c, total)

    for a in AGE_GROUPS:
        c = age_vc.get(a, 0)
        row[a] = c
        row[f"{a} (%)"] = pct(c, total)

    for p in PARTY_COLS:
        c = party_vc.get(p, 0)
        row[p] = c
        row[f"{p} (%)"] = pct(c, total)

    return row


def build_dm_row(base, grp):
    row = build_row(base, grp)

    pkr_grp = grp[
        grp["party"].astype(str).str.strip().str.upper() == "PKR"
    ]

    pkr_total = pkr_grp["nokp"].notna().sum()
    pkr_race_vc = pkr_grp["race_norm"].value_counts()

    row["PKR TOTAL"] = pkr_total

    for r in DM_EXTRA_RACES:
        c = pkr_race_vc.get(r, 0)
        row[f"PKR {r}"] = c
        row[f"PKR {r} (%)"] = pct(c, pkr_total)

    kelabu_grp = grp[
        (grp["sikap"].astype(str).str.strip().str.upper().str.contains("KELABU", na=False)) &
        (grp["umur_num"] > 30)
    ]

    kelabu_total = kelabu_grp["nokp"].notna().sum()
    kelabu_race_vc = kelabu_grp["race_norm"].value_counts()

    row["KELABU UMUR >30 TOTAL"] = kelabu_total

    for r in DM_EXTRA_RACES:
        c = kelabu_race_vc.get(r, 0)
        row[f"KELABU UMUR >30 {r}"] = c
        row[f"KELABU UMUR >30 {r} (%)"] = pct(c, kelabu_total)

    return row


def add_grand_total(df):
    if df.empty:
        return df

    id_cols = [
        "KOD PARLIMEN",
        "NAMA PARLIMEN",
        "KOD DUN",
        "NAMA DUN",
        "KOD DM",
        "NAMA DM",
    ]

    count_cols = [
        c for c in df.columns
        if c not in id_cols and "(%)" not in c
    ]

    pct_cols = [c for c in df.columns if "(%)" in c]

    gt = df[count_cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum()
    grand = gt.to_dict()

    for p in pct_cols:
        base_col = p.replace(" (%)", "")

        if base_col.startswith("PKR "):
            total_col = "PKR TOTAL"
        elif base_col.startswith("KELABU UMUR >30 "):
            total_col = "KELABU UMUR >30 TOTAL"
        else:
            total_col = "JUMLAH"

        grand[p] = pct(gt.get(base_col, 0), gt.get(total_col, 0))

    grand.update({
        "KOD PARLIMEN": "",
        "NAMA PARLIMEN": "GRAND TOTAL",
        "KOD DUN": "",
        "NAMA DUN": "",
        "KOD DM": "",
        "NAMA DM": "",
    })

    return pd.concat([df, pd.DataFrame([grand])], ignore_index=True)


def process_stats_file(file_obj):
    df = pd.read_excel(file_obj, dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]

    required_cols = [
        "nokp",
        "umur",
        "jantina",
        "kategori_kaum",
        "party",
        "sikap",
        "kod_parlimen",
        "nama_parlimen",
        "kod_dun",
        "nama_dun",
        "kod_dm",
        "nama_dm",
    ]

    missing = [c for c in required_cols if c not in df.columns]

    if missing:
        raise ValueError(f"Missing columns: {missing}")

    df["age_group"] = df["umur"].apply(get_age_group)
    df["umur_num"] = df["umur"].apply(get_age_num)
    df["race_norm"] = df["kategori_kaum"].apply(normalise_race)

    parl_rows = []
    dun_rows = []
    dm_rows = []

    for (kod_p, nama_p), grp in df.groupby(
        ["kod_parlimen", "nama_parlimen"],
        dropna=False
    ):
        parl_rows.append(build_row({
            "KOD PARLIMEN": kod_p,
            "NAMA PARLIMEN": nama_p,
            "KOD DUN": "-",
            "NAMA DUN": "(PARLIMEN TOTAL)",
            "KOD DM": "-",
            "NAMA DM": "-",
        }, grp))

    for (kod_p, nama_p, kod_d, nama_d), grp in df.groupby(
        ["kod_parlimen", "nama_parlimen", "kod_dun", "nama_dun"],
        dropna=False
    ):
        dun_rows.append(build_row({
            "KOD PARLIMEN": kod_p,
            "NAMA PARLIMEN": nama_p,
            "KOD DUN": kod_d,
            "NAMA DUN": nama_d,
            "KOD DM": "-",
            "NAMA DM": "(DUN TOTAL)",
        }, grp))

    for (kod_p, nama_p, kod_d, nama_d, kod_dm, nama_dm), grp in df.groupby(
        ["kod_parlimen", "nama_parlimen", "kod_dun", "nama_dun", "kod_dm", "nama_dm"],
        dropna=False
    ):
        dm_rows.append(build_dm_row({
            "KOD PARLIMEN": kod_p,
            "NAMA PARLIMEN": nama_p,
            "KOD DUN": kod_d,
            "NAMA DUN": nama_d,
            "KOD DM": kod_dm,
            "NAMA DM": nama_dm,
        }, grp))

    df_parl = add_grand_total(pd.DataFrame(parl_rows))
    df_dun = add_grand_total(pd.DataFrame(dun_rows))
    df_dm = add_grand_total(pd.DataFrame(dm_rows))

    return df_parl, df_dun, df_dm


def make_excel_bytes(df_parl, df_dun, df_dm):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_parl.to_excel(writer, index=False, sheet_name="BY PARLIMEN")
        df_dun.to_excel(writer, index=False, sheet_name="BY DUN")
        df_dm.to_excel(writer, index=False, sheet_name="BY DM")

        for sheet in ["BY PARLIMEN", "BY DUN", "BY DM"]:
            ws = writer.sheets[sheet]
            header = [cell.value for cell in ws[1]]

            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col),
                    default=10
                )
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(bold=False)
                    cell.border = Border()
                    cell.fill = PatternFill(fill_type=None)

                    col_name = header[cell.column - 1] if cell.column <= len(header) else ""

                    if isinstance(cell.value, (int, float)):
                        if "(%)" in str(col_name):
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = "#,##0"

    output.seek(0)
    return output.getvalue()
