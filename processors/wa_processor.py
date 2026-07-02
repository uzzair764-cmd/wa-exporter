import os
import re
import shutil
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, PatternFill


# ============================================================
# CONSTANTS
# ============================================================

ORIGINAL_KAUM_BUCKETS = ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]

KAUM_ORDER = ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]

CODE_ORDER = [
    "ML", "MP",
    "CL", "CP",
    "IL", "IP",
    "LLL", "LLP",
]

PHONE_PRIORITY_COLUMNS = ["number", "number2"]

AGE_RANGES = [
    ("18-25", "18-25", 18, 25),
    ("26-40", "26-40", 26, 40),
    ("41-60", "41-60", 41, 60),
    ("61>", "61 KE ATAS", 61, None),
]

ACTIVE_AGE_RANGES = AGE_RANGES

COLUMN_ALIASES = {
    "name": ["name", "nama", "nama penuh", "nama pemilih"],

    "nokp": [
        "nokp", "no kp", "no.kp", "no kad pengenalan",
        "ic", "no ic", "nric", "no k/p", "no kp lama"
    ],

    "jantina": ["jantina", "gender", "sex"],
    "umur": ["umur", "age"],
    "kaum_spr": ["kaum spr", "kaum_spr"],

    "kategori_kaum": [
        "kategori kaum", "kategori_kaum", "kaum", "bangsa", "race"
    ],

    "number": [
        "number", "phone", "phone 1", "no tel", "no tel 1",
        "no telefon", "nombor", "telefon", "mobile phone",
        "mobile", "mobile no", "mobile number", "no hp", "no hp 1"
    ],

    "number2": [
        "phone 2", "no tel 2", "no telefon 2", "telefon 2",
        "mobile phone 2", "mobile 2", "no hp 2"
    ],

    "kod_lokaliti": ["kod lokaliti", "kod_lokaliti", "locality code", "locality_code"],
    "nama_lokaliti": ["nama lokaliti", "nama_lokaliti", "lokaliti", "locality"],

    "kod_dm": ["kod dm", "kod_dm", "dm code", "dm_code"],
    "nama_dm": ["nama dm", "nama_dm", "dm"],

    "kod_dun": ["kod dun", "kod_dun", "dun code", "dun_code"],
    "nama_dun": ["nama dun", "nama_dun", "dun"],

    "kod_parlimen": ["kod parlimen", "kod_parlimen", "parliament code", "parliament_code"],
    "nama_parlimen": ["nama parlimen", "nama_parlimen", "parlimen", "parliament"],

    "kod_negeri": ["kod negeri", "kod_negeri", "state code", "state_code"],
    "nama_negeri": ["nama negeri", "nama_negeri", "negeri", "state"],

    "sikap": ["sikap"],
    "party": ["party", "parti"],
}


# ============================================================
# BASIC HELPERS
# ============================================================

def emit(progress_callback, message, progress_value=None):
    if progress_callback:
        progress_callback(message, progress_value)


def normalize_key(text):
    text = str(text)
    text = text.replace("\xa0", " ")
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = text.replace("_", " ")
    text = text.strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


ALIAS_LOOKUP = {}

for canonical, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        ALIAS_LOOKUP[normalize_key(alias)] = canonical


def sanitize(text):
    text = str(text).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return re.sub(r'[\\/*?:"<>|]', "_", text)


def format_count(n):
    return f"{int(n):,}"


def make_unique_columns(cols):
    seen = {}
    final_cols = []

    for col in cols:
        norm = normalize_key(col)

        if norm in ALIAS_LOOKUP:
            new_col = ALIAS_LOOKUP[norm]
        else:
            new_col = str(col).strip()
            new_col = new_col.replace("\xa0", " ")
            new_col = re.sub(r"\s+", " ", new_col)

            if new_col == "" or new_col.upper().startswith("UNNAMED"):
                new_col = "UNNAMED"

        if new_col not in seen:
            seen[new_col] = 0
            final_cols.append(new_col)
        else:
            seen[new_col] += 1
            final_cols.append(f"{new_col}__{seen[new_col]}")

    return final_cols


def detect_header_row(raw_df):
    best_row = None
    best_score = 0
    max_scan = min(100, len(raw_df))

    for i in range(max_scan):
        row_values = [normalize_key(x) for x in raw_df.iloc[i].tolist()]
        score = sum(1 for x in row_values if x in ALIAS_LOOKUP)

        if score > best_score:
            best_score = score
            best_row = i

    if best_score >= 3:
        return best_row

    return None


def read_excel_smart(file_name, read_all_sheets=True, progress_callback=None):
    all_sheets = []

    xls = pd.ExcelFile(file_name)
    sheet_names = xls.sheet_names if read_all_sheets else [xls.sheet_names[0]]

    emit(progress_callback, f"Reading: {os.path.basename(file_name)}", 8)

    for sheet_name in sheet_names:
        raw = pd.read_excel(
            file_name,
            sheet_name=sheet_name,
            header=None,
            dtype=str
        ).fillna("")

        header_row = detect_header_row(raw)

        if header_row is None:
            continue

        cols = raw.iloc[header_row].tolist()
        cols = make_unique_columns(cols)

        data = raw.iloc[header_row + 1:].copy()
        data.columns = cols
        data = data.fillna("")

        data = data[
            data.astype(str).apply(
                lambda row: "".join(row.values).strip() != "",
                axis=1
            )
        ].copy()

        if data.empty:
            continue

        data["source_file"] = os.path.basename(file_name)
        data["source_sheet"] = sheet_name

        all_sheets.append(data)

    if not all_sheets:
        raise ValueError(f"No usable sheet found in {os.path.basename(file_name)}")

    return pd.concat(all_sheets, ignore_index=True)


# ============================================================
# CLEANING HELPERS
# ============================================================

def clean_phone(value):
    phone = re.sub(r"\D", "", str(value).strip())

    if phone.startswith("60") and len(phone) in [11, 12]:
        phone = "0" + phone[2:]

    if phone == "":
        return None

    if "0000" in phone:
        return None

    if not phone.startswith("01"):
        return None

    if not phone.isdigit():
        return None

    if not (10 <= len(phone) <= 11):
        return None

    return phone


def choose_one_phone(row):
    for col in PHONE_PRIORITY_COLUMNS:
        if col in row:
            phone = clean_phone(row.get(col, ""))

            if phone is not None:
                return phone

    return None


def normalize_filter_value(x):
    x = str(x).strip().upper()
    x = x.replace("_", " ")
    x = re.sub(r"\s+", " ", x)

    if x in ["LAIN LAIN", "LAIN-LAIN", "LAINLAIN"]:
        return "LAIN-LAIN"

    if x == "ORANG ASLI":
        return "LAIN-LAIN"

    if "BUMI" in x:
        return "LAIN-LAIN"

    if x not in ["MELAYU", "CINA", "INDIA", "LAIN-LAIN"]:
        return "LAIN-LAIN"

    return x


def clean_kaum(value):
    k = str(value).strip().upper()
    k = k.replace("_", " ")
    k = re.sub(r"\s+", " ", k)

    if k == "MELAYU":
        return "MELAYU"

    elif k == "CINA":
        return "CINA"

    elif k == "INDIA":
        return "INDIA"

    else:
        return "LAIN-LAIN"


def normalize_sikap(value):
    s = str(value).strip().upper()

    s = s.replace("\xa0", " ")
    s = s.replace("_", " ")
    s = s.replace("–", "-")
    s = s.replace("—", "-")
    s = re.sub(r"\s+", " ", s)

    compact = re.sub(r"[^A-Z]", "", s)

    if "KELABU" in compact:
        return "KELABU"

    if "HITAM" in compact:
        return "HITAM"

    if "PUTIH" in compact:
        return "PUTIH"

    return s


# ============================================================
# WA FORMAT HELPERS
# ============================================================

def get_code(kategori_kaum, gender):
    k = clean_kaum(kategori_kaum)
    g = str(gender).strip().upper()

    if k == "MELAYU":
        return "ML" if g == "L" else "MP"

    elif k == "CINA":
        return "CL" if g == "L" else "CP"

    elif k == "INDIA":
        return "IL" if g == "L" else "IP"

    else:
        return "LLL" if g == "L" else "LLP"


def format_code_7digit(raw_code):
    raw = str(raw_code).strip()

    if "/" in raw:
        parts = raw.split("/")

        if len(parts) >= 3:
            parlimen = re.sub(r"\D", "", parts[0]).zfill(3)
            dun = re.sub(r"\D", "", parts[1]).zfill(2)
            last_raw = re.sub(r"\D", "", parts[2])
            last = f"{int(last_raw):02d}" if last_raw else "00"

            return f".{parlimen}.{dun}.{last}"

    digits = re.sub(r"\D", "", raw)

    if digits == "":
        return ""

    digits = digits.zfill(7)

    parlimen = digits[:3]
    dun = digits[3:5]
    last_raw = digits[5:7]
    last = f"{int(last_raw):02d}" if last_raw else "00"

    return f".{parlimen}.{dun}.{last}"


def format_first_name(row):
    return format_code_7digit(row.get("kod_dm", ""))


def build_wa_df(group):
    return pd.DataFrame({
        "Title": group["number"].astype(str),
        "First Name": group["First Name"].astype(str),
        "Middle Name": "",
        "Last Name": group["Last Name"].astype(str),
        "Mobile Phone": group["number"].astype(str),
    })


# ============================================================
# LABEL HELPERS
# ============================================================

def assign_age_range(age, age_ranges=None):
    if age_ranges is None:
        age_ranges = ACTIVE_AGE_RANGES

    try:
        age = int(float(str(age).strip()))
    except:
        return None

    for txt_label, file_label, min_age, max_age in age_ranges:
        if max_age is None:
            if age >= min_age:
                return txt_label
        else:
            if min_age <= age <= max_age:
                return txt_label

    return None


def normalize_parlimen_code(code):
    raw = str(code).strip().upper()
    num = re.sub(r"\D", "", raw)

    if num == "":
        return "P000"

    return f"P{int(num):03d}"


def extract_dun_num(kod_dun):
    raw = str(kod_dun).strip().upper()
    num = re.sub(r"\D", "", raw)

    if num == "":
        return "00"

    if len(num) >= 2:
        return f"{int(num[-2:]):02d}"

    return f"{int(num):02d}"


def extract_dm_num(kod_dm):
    raw = str(kod_dm).strip()

    if "/" in raw:
        parts = raw.split("/")

        if len(parts) >= 3:
            dm_raw = re.sub(r"\D", "", parts[2])
            return f"{int(dm_raw):02d}" if dm_raw else "00"

    digits = re.sub(r"\D", "", raw)

    if digits == "":
        return "00"

    digits = digits.zfill(7)
    dm_raw = digits[-2:]

    return f"{int(dm_raw):02d}" if dm_raw else "00"


def code_sort_num(text):
    num = re.sub(r"\D", "", str(text))
    return int(num) if num else 0


def get_file_label_from_age(txt_label):
    active_age_ranges = globals().get("ACTIVE_AGE_RANGES", AGE_RANGES)

    for t, file_label, min_age, max_age in active_age_ranges:
        if t == txt_label:
            return file_label

    return txt_label


def safe_file_label(label):
    label = str(label)

    if label in [x[0] for x in globals().get("ACTIVE_AGE_RANGES", AGE_RANGES)]:
        label = get_file_label_from_age(label)

    return sanitize(label)


def build_labels(df):
    if "kod_parlimen" in df.columns:
        df["__PARLIMEN_LABEL"] = df.apply(
            lambda x: f"{normalize_parlimen_code(x.get('kod_parlimen', ''))} {sanitize(x.get('nama_parlimen', ''))}",
            axis=1
        )
        df["__PARLIMEN_SORT"] = df["kod_parlimen"].apply(code_sort_num)

    if "kod_dun" in df.columns:
        df["__DUN_LABEL"] = df.apply(
            lambda x: f"N.{extract_dun_num(x.get('kod_dun', ''))} {sanitize(x.get('nama_dun', ''))}",
            axis=1
        )
        df["__DUN_SORT"] = df["kod_dun"].apply(code_sort_num)

    if "kod_dm" in df.columns:
        df["__DM_LABEL"] = df.apply(
            lambda x: f"DM{extract_dm_num(x.get('kod_dm', ''))} {sanitize(x.get('nama_dm', ''))}",
            axis=1
        )
        df["__DM_SORT"] = df["kod_dm"].apply(code_sort_num)

    if "kaum_clean" in df.columns:
        df["__KAUM_LABEL"] = df["kaum_clean"]

        kaum_order = {
            "MELAYU": 1,
            "CINA": 2,
            "INDIA": 3,
            "LAIN-LAIN": 4,
        }

        df["__KAUM_SORT"] = df["kaum_clean"].map(kaum_order).fillna(99).astype(int)

    if "umur" in df.columns:
        active_age_ranges = globals().get("ACTIVE_AGE_RANGES", AGE_RANGES)

        df["__AGE_LABEL"] = df["umur"].apply(
            lambda x: assign_age_range(x, active_age_ranges)
        )

        age_order = {
            txt_label: i
            for i, (txt_label, file_label, min_age, max_age) in enumerate(active_age_ranges)
        }

        df["__AGE_SORT"] = df["__AGE_LABEL"].map(age_order).fillna(999).astype(int)

    return df


def get_group_columns(group_levels):
    label_cols = []
    sort_cols = []

    for level in group_levels:
        if level == "PARLIMEN":
            label_cols.append("__PARLIMEN_LABEL")
            sort_cols.append("__PARLIMEN_SORT")

        elif level == "DUN":
            label_cols.append("__DUN_LABEL")
            sort_cols.append("__DUN_SORT")

        elif level == "DM":
            label_cols.append("__DM_LABEL")
            sort_cols.append("__DM_SORT")

        elif level == "KAUM":
            label_cols.append("__KAUM_LABEL")
            sort_cols.append("__KAUM_SORT")

        elif level == "AGE":
            label_cols.append("__AGE_LABEL")
            sort_cols.append("__AGE_SORT")

    return label_cols, sort_cols


# ============================================================
# VALIDATION
# ============================================================

def get_required_columns(config):
    req = set()

    group_levels = config["group_levels"]
    sikap_filter = config["sikap_filter"]
    party_filter = config["party_filter"]
    age_filter = config["age_filter"]

    req.add("number")
    req.add("kod_dm")
    req.add("kategori_kaum")
    req.add("jantina")

    for level in group_levels:
        if level == "PARLIMEN":
            req.update(["kod_parlimen", "nama_parlimen"])

        elif level == "DUN":
            req.update(["kod_dun", "nama_dun"])

        elif level == "DM":
            req.update(["kod_dm", "nama_dm"])

        elif level == "KAUM":
            req.add("kategori_kaum")

        elif level == "AGE":
            req.add("umur")

    if sikap_filter:
        req.add("sikap")

    if party_filter:
        req.add("party")

    if age_filter is not None:
        req.add("umur")

    return sorted(req)


def validate_required_columns(df, config):
    req = get_required_columns(config)
    missing = []

    for col in req:
        if col == "number":
            if "number" not in df.columns and "number2" not in df.columns:
                missing.append("number / PHONE 1")
        else:
            if col not in df.columns:
                missing.append(col)

    if missing:
        raise ValueError(
            "Missing columns after header cleanup: "
            + str(missing)
            + "\n\nAvailable columns:\n"
            + str(list(df.columns))
        )


# ============================================================
# FILE WRITER
# ============================================================

def save_xlsx_no_style(df, path):
    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(bold=False)
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="left", vertical="top")

    wb.save(path)


def get_split_items(group, split_col):
    items = []

    if split_col == "kaum_clean":
        for key in KAUM_ORDER:
            sub = group[group["kaum_clean"] == key].copy()
            if not sub.empty:
                items.append((key, sub))

        return items

    if split_col == "code":
        for key in CODE_ORDER:
            sub = group[group["code"] == key].copy()
            if not sub.empty:
                items.append((key, sub))

        return items

    for key, sub in group.groupby(split_col, dropna=False, sort=False):
        items.append((key, sub.copy()))

    return items


def build_summary_output_lines(output_lines):
    grouped_summary = {}
    no_group_lines = []

    for line in output_lines:
        if not line.strip():
            continue

        if " / " in line and " = " in line and not line.startswith("  "):
            left, count = line.split(" = ", 1)
            parts = left.split(" / ")

            if len(parts) >= 2:
                parent = parts[0]
                child = parts[-1]
                grouped_summary.setdefault(parent, []).append({
                    "title": child,
                    "count": count,
                    "children": []
                })

        elif line.startswith("  ") and grouped_summary:
            last_parent = list(grouped_summary.keys())[-1]

            if grouped_summary[last_parent]:
                grouped_summary[last_parent][-1]["children"].append(line)

        elif " = " in line:
            no_group_lines.append(line)

    final_lines = []

    if grouped_summary:
        for parent, items in grouped_summary.items():
            final_lines.append(parent)
            final_lines.append("")

            for idx, item in enumerate(items, start=1):
                final_lines.append(f"{idx}. {item['title']} = {item['count']}")

                for child in item["children"]:
                    final_lines.append(child)

                final_lines.append("")

    else:
        final_lines.extend(no_group_lines)

    return final_lines


# ============================================================
# ZIP NAME HELPER
# ============================================================

def build_auto_zip_name(file_paths, df, input_level):
    if len(file_paths) == 1:
        original_name = os.path.splitext(os.path.basename(file_paths[0]))[0]
        return sanitize(f"WA {original_name}")

    negeri_list = []

    if "nama_negeri" in df.columns:
        negeri_list = (
            df["nama_negeri"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .replace("", pd.NA)
            .dropna()
            .unique()
            .tolist()
        )

    if len(negeri_list) == 1:
        if str(input_level).strip().upper() == "PARLIMEN":
            return sanitize(f"WA PARLIMEN {negeri_list[0]}")
        return sanitize(f"WA DUN {negeri_list[0]}")

    if str(input_level).strip().upper() == "PARLIMEN":
        return sanitize("WA MULTIPLE PARLIMEN")

    return sanitize("WA MULTIPLE DUN")


# ============================================================
# MAIN EXPORT FUNCTION
# ============================================================

def run_export(file_paths, config, progress_callback=None):
    global ACTIVE_AGE_RANGES

    output_dir = config["output_dir"]

    input_level = config["input_level"]
    group_levels = config["group_levels"]
    last_group_as_folder = config["last_group_as_folder"]
    split_by_kaum = config["split_by_kaum"]
    split_by_gender_code = config["split_by_gender_code"]
    keep_kaum = config["keep_kaum"]
    sikap_filter = config["sikap_filter"]
    party_filter = config["party_filter"]
    age_filter = config["age_filter"]
    custom_age_ranges = config.get("custom_age_ranges", [])
    dedup_by_phone = config["dedup_by_phone"]
    dedup_by_nokp = config["dedup_by_nokp"]
    read_all_sheets = config["read_all_sheets"]
    create_empty_files = config["create_empty_files"]

    if custom_age_ranges:
        ACTIVE_AGE_RANGES = custom_age_ranges
    else:
        ACTIVE_AGE_RANGES = AGE_RANGES

    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)

    os.makedirs(output_dir, exist_ok=True)

    emit(progress_callback, "Reading uploaded Excel file(s)...", 5)

    all_data = []

    for file_name in file_paths:
        df_file = read_excel_smart(
            file_name=file_name,
            read_all_sheets=read_all_sheets,
            progress_callback=progress_callback
        )
        all_data.append(df_file)

    if not all_data:
        raise ValueError("No uploaded data found.")

    df = pd.concat(all_data, ignore_index=True)

    original_rows = len(df)

    if "number2" not in df.columns:
        df["number2"] = ""

    validate_required_columns(df, config)

    emit(progress_callback, "Cleaning phone numbers...", 15)

    df["number"] = df.apply(choose_one_phone, axis=1)

    before_phone = len(df)
    df = df[df["number"].notna()].copy()
    after_phone = len(df)

    if df.empty:
        raise ValueError("No rows left after number cleaning.")

    if dedup_by_nokp and "nokp" in df.columns:
        emit(progress_callback, "Removing duplicate NO KP...", 25)

        df["nokp_clean"] = df["nokp"].astype(str).str.replace(r"\D", "", regex=True).str.strip()

        df_valid_nokp = df[df["nokp_clean"] != ""].drop_duplicates(
            subset=["nokp_clean"],
            keep="first"
        )

        df_blank_nokp = df[df["nokp_clean"] == ""]

        df = pd.concat([df_valid_nokp, df_blank_nokp], ignore_index=True)

    if dedup_by_phone:
        emit(progress_callback, "Removing duplicate phone numbers...", 30)
        before_phone_dedup = len(df)
        df = df.drop_duplicates(subset=["number"], keep="first").copy()
        after_phone_dedup = len(df)
    else:
        before_phone_dedup = len(df)
        after_phone_dedup = len(df)

    if df.empty:
        raise ValueError("No rows left after duplicate cleaning.")

    emit(progress_callback, "Cleaning kaum...", 40)

    df["kaum_clean"] = df["kategori_kaum"].apply(clean_kaum)

    before_kaum = len(df)

    if keep_kaum != ["ALL"]:
        keep_set = set(normalize_filter_value(x) for x in keep_kaum)
        df = df[df["kaum_clean"].isin(keep_set)].copy()

    after_kaum = len(df)

    if df.empty:
        raise ValueError("No rows left after kaum filtering.")

    if "sikap" in df.columns:
        df["sikap_clean"] = df["sikap"].apply(normalize_sikap)
    else:
        df["sikap_clean"] = ""

    if sikap_filter:
        emit(progress_callback, "Filtering sikap...", 50)

        sikap_set = set(normalize_sikap(x) for x in sikap_filter)
        df = df[df["sikap_clean"].isin(sikap_set)].copy()

    if df.empty:
        raise ValueError("No rows left after sikap filtering.")

    if party_filter:
        emit(progress_callback, "Filtering party...", 55)

        party_set = set(str(x).strip().upper() for x in party_filter)

        df["party_clean"] = df["party"].astype(str).str.strip().str.upper()
        df = df[df["party_clean"].isin(party_set)].copy()

    if df.empty:
        raise ValueError("No rows left after party filtering.")

    if age_filter is not None:
        emit(progress_callback, "Filtering age group(s)...", 60)

        df["umur_num"] = pd.to_numeric(df["umur"], errors="coerce")

        if age_filter == "CUSTOM_AGE_GROUPS" and custom_age_ranges:
            mask = False

            for label, file_label, min_age, max_age in custom_age_ranges:
                if max_age is None:
                    current_mask = df["umur_num"] >= min_age
                else:
                    current_mask = (
                        (df["umur_num"] >= min_age) &
                        (df["umur_num"] <= max_age)
                    )

                mask = mask | current_mask

            df = df[mask].copy()

        else:
            min_age, max_age = age_filter

            if max_age is None:
                df = df[df["umur_num"] >= min_age].copy()

            else:
                df = df[
                    (df["umur_num"] >= min_age) &
                    (df["umur_num"] <= max_age)
                ].copy()

    if df.empty:
        raise ValueError("No rows left after age filtering.")

    emit(progress_callback, "Preparing WhatsApp columns...", 70)

    df["code"] = df.apply(
        lambda x: get_code(x.get("kategori_kaum", ""), x.get("jantina", "")),
        axis=1
    )

    df["First Name"] = df.apply(format_first_name, axis=1)

    df = df[df["First Name"].astype(str).str.strip() != ""].copy()

    df["Last Name"] = df["code"]

    if df.empty:
        raise ValueError("No rows left after WhatsApp formatting.")

    emit(progress_callback, "Preparing output groups...", 75)

    df = build_labels(df)

    if "AGE" in group_levels:
        df = df[df["__AGE_LABEL"].notna()].copy()

    if df.empty:
        raise ValueError("No rows left after grouping setup.")

    emit(progress_callback, "Writing Excel files...", 82)

    output_lines = []

    label_cols, sort_cols = get_group_columns(group_levels)

    sort_existing = [c for c in sort_cols if c in df.columns]

    if sort_existing:
        df = df.sort_values(sort_existing).copy()

    split_col = None

    if split_by_gender_code:
        split_col = "code"

    elif split_by_kaum:
        split_col = "kaum_clean"

    total_files_created = 0

    if not label_cols:
        base_folder = output_dir
        os.makedirs(base_folder, exist_ok=True)

        if split_col:
            split_items = get_split_items(df, split_col)

            for split_value, split_group in split_items:
                if split_group.empty and not create_empty_files:
                    continue

                export_df = build_wa_df(split_group)

                out_file = f"{safe_file_label(split_value)}.xlsx"
                out_path = os.path.join(base_folder, out_file)

                save_xlsx_no_style(export_df, out_path)

                total_files_created += 1
                output_lines.append(f"{safe_file_label(split_value)} = {format_count(len(split_group))}")

        else:
            export_df = build_wa_df(df)

            out_file = "OUTPUT.xlsx"
            out_path = os.path.join(base_folder, out_file)

            save_xlsx_no_style(export_df, out_path)

            total_files_created += 1
            output_lines.append(f"OUTPUT = {format_count(len(df))}")

    else:
        grouped = df.groupby(label_cols, dropna=False, sort=False)

        for group_key, group in grouped:
            if not isinstance(group_key, tuple):
                group_key = (group_key,)

            group_labels = [safe_file_label(x) for x in group_key]

            if len(group_labels) == 1:
                # Single-level outputs like:
                #   DUN > DUN.xlsx
                #   PARLIMEN > PARLIMEN.xlsx
                # should be written directly under voter_outputs/, not inside a folder.
                # Split modes like DUN > KAUM still use a DUN folder because split_col is active.
                if last_group_as_folder and split_col:
                    folder_parts = group_labels
                    file_base = group_labels[-1]

                else:
                    folder_parts = []
                    file_base = group_labels[-1]

            else:
                if last_group_as_folder:
                    folder_parts = group_labels
                    file_base = group_labels[-1]

                else:
                    folder_parts = group_labels[:-1]
                    file_base = group_labels[-1]

            out_folder = os.path.join(output_dir, *folder_parts)
            os.makedirs(out_folder, exist_ok=True)

            display_path = " / ".join(group_labels)

            if split_col:
                split_items = get_split_items(group, split_col)
                output_lines.append(f"{display_path} = {format_count(len(group))}")

                for split_value, split_group in split_items:
                    if split_group.empty and not create_empty_files:
                        continue

                    split_label = safe_file_label(split_value)
                    export_df = build_wa_df(split_group)

                    out_file = f"{file_base} {split_label}.xlsx"
                    out_file = sanitize(out_file)
                    out_path = os.path.join(out_folder, out_file)

                    save_xlsx_no_style(export_df, out_path)

                    total_files_created += 1
                    output_lines.append(f"  {split_label} = {format_count(len(split_group))}")

            else:
                if group.empty and not create_empty_files:
                    continue

                export_df = build_wa_df(group)

                out_file = f"{file_base}.xlsx"
                out_file = sanitize(out_file)
                out_path = os.path.join(out_folder, out_file)

                save_xlsx_no_style(export_df, out_path)

                total_files_created += 1
                output_lines.append(f"{display_path} = {format_count(len(group))}")

            output_lines.append("")

    emit(progress_callback, "Creating summary...", 92)

    summary_output_lines = build_summary_output_lines(output_lines)

    summary_text = "\n".join(summary_output_lines).strip() + "\n"

    summary_path = os.path.join(output_dir, "SUMMARY.txt")

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary_text)

    filter_log_lines = [
        "FILTER LOG",
        "",
        f"TOTAL INPUT ROWS = {format_count(original_rows)}",
        f"FINAL ROWS = {format_count(len(df))}",
        f"EXCEL FILES CREATED = {format_count(total_files_created)}",
        "",
        "FILTERS",
        f"INPUT LEVEL = {input_level}",
        f"KAUM = {keep_kaum}",
        f"SIKAP = {sikap_filter if sikap_filter else 'NO FILTER'}",
        f"PARTY = {party_filter if party_filter else 'NO FILTER'}",
        f"AGE GROUPS = {custom_age_ranges if custom_age_ranges else 'NO FILTER'}",
        "",
        "CLEANING",
        f"VALID PHONE ROWS = {format_count(after_phone)}",
        f"DUPLICATE PHONES REMOVED = {format_count(before_phone_dedup - after_phone_dedup)}",
    ]

    filter_log_text = "\n".join(filter_log_lines).strip() + "\n"

    filter_log_path = os.path.join(output_dir, "FILTER LOG.txt")

    with open(filter_log_path, "w", encoding="utf-8") as f:
        f.write(filter_log_text)

    emit(progress_callback, "Creating ZIP...", 96)

    zip_filename = build_auto_zip_name(
        file_paths=file_paths,
        df=df,
        input_level=input_level
    )

    zip_base = os.path.join(os.path.dirname(output_dir), zip_filename)
    zip_path = shutil.make_archive(zip_base, "zip", output_dir)

    emit(progress_callback, "Done.", 100)

    return {
        "final_rows": int(len(df)),
        "files_created": int(total_files_created),
        "zip_path": zip_path,
        "summary_path": summary_path,
        "summary_text": summary_text,
    }
